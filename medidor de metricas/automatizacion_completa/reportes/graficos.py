"""Generación de gráficos para el reporte Excel."""

from openpyxl.chart import BarChart, Reference


def generar_graficos(libro, configuracion):
    """Regenera la hoja de gráficos según la configuración recibida.

    Args:
        libro: Libro de Excel que contiene las hojas de datos.
        configuracion: Secuencia de configuraciones para cada gráfico.
    """
    nombre_hoja = "Graficos"
    if nombre_hoja in libro.sheetnames:
        del libro[nombre_hoja]
    hoja_graficos = libro.create_sheet(nombre_hoja)

    for nombre_datos, columna, titulo, eje_y, posicion in configuracion:
        hoja_datos = libro[nombre_datos]
        if hoja_datos.max_row >= 2:
            agregar_grafico_barras(
                hoja_graficos, hoja_datos, columna, titulo, eje_y, posicion
            )


def agregar_grafico_barras(hoja_destino, hoja_datos, columna, titulo, eje_y, posicion):
    """Agrega un gráfico de barras usando el código como categoría.

    Args:
        hoja_destino: Hoja en la que se inserta el gráfico.
        hoja_datos: Hoja que proporciona categorías y valores.
        columna: Número de la columna que contiene los valores.
        titulo: Título del gráfico.
        eje_y: Etiqueta del eje vertical.
        posicion: Celda de anclaje del gráfico en la hoja de destino.
    """
    grafico = BarChart()
    grafico.title = titulo
    grafico.y_axis.title = eje_y
    grafico.x_axis.title = "Proyecto"
    datos = Reference(
        hoja_datos,
        min_col=columna,
        min_row=1,
        max_row=hoja_datos.max_row,
    )
    categorias = Reference(
        hoja_datos,
        min_col=1,
        min_row=2,
        max_row=hoja_datos.max_row,
    )
    grafico.add_data(datos, titles_from_data=True)
    grafico.set_categories(categorias)
    hoja_destino.add_chart(grafico, posicion)
