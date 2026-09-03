"""Creación y actualización del libro Excel de resultados."""

from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

from reportes.graficos import generar_graficos


HOJA_RESUMEN = "Resumen"
HOJA_COMPLEJIDAD = "Complejidad"
HOJA_MANTENIBILIDAD = "Mantenibilidad"
HOJA_BUGS_SMELLS = "Bugs_Smells"
HOJA_ERRORES = "Errores"
HOJA_CONCURRENCIA = "Concurrencia"

ENCABEZADOS_RESUMEN = [
    "Código",
    "Nombre del proyecto",
    "Herramienta IA",
    "Modelo IA",
    "Lenguaje",
    "CCN promedio",
    "Nivel de CC",
    "MI",
    "Nivel de MI",
    "Issues/KLOC",
    "ISI",
    "Interpretación de Issue",
    "Promedio concurrencia",
    "Interpretación concurrencia",
]

ENCABEZADOS_COMPLEJIDAD = [
    "Código",
    "Cantidad de funciones",
    "CCN Total",
    "CCN promedio",
    "NLOC total",
    "Nivel de CC",
    "Interpretación CC",
    "NLOC promedio",
]

ENCABEZADOS_MANTENIBILIDAD = [
    "Código",
    "NLOC MI",
    "Cantidad de funciones MI",
    "Tokens código",
    "MI",
    "Nivel de MI",
    "Interpretación MI",
]

ENCABEZADOS_BUGS_SMELLS = [
    "Código",
    "Analizador",
    "Total de issues",
    "Issues/KLOC",
    "ISI",
    "Nivel de ISI",
    "Interpretación de ISI",
    "Observacion",
    "Cant. severidad alta",
    "Cant. severidad media",
    "Cant. severidad baja",
    "Reglas incumplidas",
]

ENCABEZADOS_ERRORES = [
    "Código",
    "Nombre del proyecto",
    "Error",
]

ENCABEZADOS_CONCURRENCIA = [
    "Código",
    "Sincronización correcta",
    "Ausencia de deadlocks",
    "Ausencia de condición de carrera",
    "Uso correcto de exclusión mutua",
    "Promedio concurrencia",
    "Interpretación concurrencia",
]

HOJAS_REPORTE = {
    HOJA_RESUMEN: ENCABEZADOS_RESUMEN,
    HOJA_COMPLEJIDAD: ENCABEZADOS_COMPLEJIDAD,
    HOJA_MANTENIBILIDAD: ENCABEZADOS_MANTENIBILIDAD,
    HOJA_BUGS_SMELLS: ENCABEZADOS_BUGS_SMELLS,
    HOJA_ERRORES: ENCABEZADOS_ERRORES,
    HOJA_CONCURRENCIA: ENCABEZADOS_CONCURRENCIA,
}

CONFIGURACION_GRAFICOS = [
    (HOJA_RESUMEN, 6, "CCN promedio por proyecto", "CCN promedio", "A1"),
    (HOJA_COMPLEJIDAD, 5, "NLOC total por proyecto", "NLOC total", "A18"),
    (HOJA_MANTENIBILIDAD, 5, "MI por proyecto", "Índice de mantenibilidad", "A35"),
    (HOJA_BUGS_SMELLS, 5, "ISI por proyecto", "Índice de severidad de issues", "A52"),
    (HOJA_RESUMEN, 13, "Promedio de concurrencia por proyecto", "Promedio concurrencia", "A69"),
]


def crear_o_abrir_excel(archivo_excel):
    """Abre un libro existente o crea uno con las hojas requeridas.

    Args:
        archivo_excel: Ruta del libro de resultados.

    Returns:
        El libro preparado para recibir datos.
    """
    if Path(archivo_excel).exists():
        libro = load_workbook(archivo_excel)
    else:
        libro = Workbook()
        libro.remove(libro.active)

    crear_hojas_si_no_existen(libro)
    return libro


def asegurar_encabezados(hoja, encabezados):
    """Crea o actualiza encabezados sin borrar datos existentes.

    Args:
        hoja: Hoja cuyos encabezados deben actualizarse.
        encabezados: Encabezados que deben estar presentes.
    """
    if hoja.max_row == 1 and all(celda.value is None for celda in hoja[1]):
        hoja.append(encabezados)
        hoja.delete_rows(1)
        return

    encabezados_actuales = [hoja.cell(row=1, column=col).value for col in range(1, hoja.max_column + 1)]

    for encabezado in encabezados:
        if encabezado not in encabezados_actuales:
            hoja.cell(row=1, column=hoja.max_column + 1).value = encabezado
            encabezados_actuales.append(encabezado)



def eliminar_columnas_obsoletas(hoja, encabezados_validos):
    """Elimina las columnas ajenas al diseño esperado.

    Args:
        hoja: Hoja que se desea depurar.
        encabezados_validos: Encabezados que deben conservarse.
    """
    encabezados_validos = set(encabezados_validos)

    for columna in range(hoja.max_column, 0, -1):
        encabezado = hoja.cell(row=1, column=columna).value
        if encabezado is not None and encabezado not in encabezados_validos:
            hoja.delete_cols(columna)

def crear_hojas_si_no_existen(libro):
    """Crea y prepara todas las hojas requeridas por el reporte.

    Args:
        libro: Libro de Excel que se desea preparar.
    """
    for nombre, encabezados in HOJAS_REPORTE.items():
        if nombre not in libro.sheetnames:
            libro.create_sheet(nombre).append(encabezados)
            continue

        hoja = libro[nombre]
        asegurar_encabezados(hoja, encabezados)

        if nombre == HOJA_BUGS_SMELLS:
            eliminar_columnas_obsoletas(hoja, encabezados)


def aplicar_estilos_basicos(libro):
    """Aplica formato a encabezados y ajusta anchos de columnas.

    Args:
        libro: Libro de Excel que se desea formatear.
    """
    for hoja in libro.worksheets:
        for celda in hoja[1]:
            celda.font = Font(bold=True)
            celda.fill = PatternFill("solid", fgColor="D9EAF7")
            celda.alignment = Alignment(horizontal="center")

        for columna in hoja.columns:
            max_largo = 0
            letra = columna[0].column_letter

            for celda in columna:
                if celda.value is not None:
                    max_largo = max(max_largo, len(str(celda.value)))

            hoja.column_dimensions[letra].width = min(max_largo + 3, 45)


def buscar_fila_por_codigo(hoja, codigo):
    """Busca una fila por el código almacenado en la primera columna.

    Args:
        hoja: Hoja donde se realiza la búsqueda.
        codigo: Código de proyecto buscado.

    Returns:
        El número de fila encontrado o ``None``.
    """
    for fila in range(2, hoja.max_row + 1):
        if hoja.cell(row=fila, column=1).value == codigo:
            return fila

    return None


def escribir_o_actualizar_fila(hoja, codigo, valores):
    """Agrega una fila o actualiza la que corresponde a un código.

    Args:
        hoja: Hoja que se desea modificar.
        codigo: Código usado como clave de la fila.
        valores: Valores que deben escribirse.
    """
    fila_existente = buscar_fila_por_codigo(hoja, codigo)

    if fila_existente is None:
        hoja.append(valores)
    else:
        for columna, valor in enumerate(valores, start=1):
            hoja.cell(row=fila_existente, column=columna).value = valor


def escribir_valores_por_encabezado(hoja, fila, valores):
    """Escribe varios valores en columnas identificadas por sus encabezados.

    Args:
        hoja: Hoja que se desea modificar.
        fila: Número de fila de destino.
        valores: Pares formados por encabezado y valor.
    """
    encabezados = obtener_mapa_encabezados(hoja)

    for encabezado, valor in valores.items():
        columna = encabezados.get(encabezado)
        if columna is not None:
            hoja.cell(row=fila, column=columna).value = valor


def obtener_mapa_encabezados(hoja):
    """Crea un mapa entre encabezados y números de columna.

    Args:
        hoja: Hoja que se desea inspeccionar.

    Returns:
        Diccionario de encabezados a columnas.
    """
    return {
        hoja.cell(row=1, column=columna).value: columna
        for columna in range(1, hoja.max_column + 1)
    }


def finalizar_libro(libro, archivo_excel, incluir_graficos=False):
    """Aplica las tareas finales y guarda el libro una sola vez.

    Args:
        libro: Libro de Excel que se desea finalizar.
        archivo_excel: Ruta en la que se guarda el libro.
        incluir_graficos: Indica si deben regenerarse los gráficos.
    """
    aplicar_estilos_basicos(libro)

    if incluir_graficos:
        generar_graficos(libro, CONFIGURACION_GRAFICOS)

    ruta_salida = Path(archivo_excel)
    ruta_salida.parent.mkdir(parents=True, exist_ok=True)
    ruta_temporal = ruta_salida.with_name(f".{ruta_salida.stem}.tmp{ruta_salida.suffix}")
    libro.save(ruta_temporal)
    ruta_temporal.replace(ruta_salida)


def guardar_resultado_excel(
    libro,
    proyecto,
    metricas_cc,
    metricas_mi,
    metricas_bugs_smells=None,
    metricas_concurrencia=None,
):
    """Guarda todas las métricas de un proyecto en Excel.

    Args:
        libro: Libro de Excel en el que se escriben las métricas.
        proyecto: Proyecto analizado.
        metricas_cc: Métricas de complejidad.
        metricas_mi: Métricas de mantenibilidad.
        metricas_bugs_smells: Métricas de incidencias, si existen.
        metricas_concurrencia: Métricas de concurrencia, si existen.
    """
    hoja_resumen = libro[HOJA_RESUMEN]
    hoja_complejidad = libro[HOJA_COMPLEJIDAD]
    hoja_mantenibilidad = libro[HOJA_MANTENIBILIDAD]
    hoja_bugs_smells = libro[HOJA_BUGS_SMELLS]
    hoja_concurrencia = libro[HOJA_CONCURRENCIA]

    escribir_hoja_resumen(hoja_resumen, proyecto, metricas_cc, metricas_mi, metricas_bugs_smells, metricas_concurrencia)
    
    escribir_hoja_complejidad(hoja_complejidad, proyecto.codigo, metricas_cc)
    escribir_hoja_mantenibilidad(hoja_mantenibilidad, proyecto.codigo, metricas_mi)
    escribir_hoja_bugs_smells(hoja_bugs_smells, proyecto, metricas_bugs_smells)

    escribir_hoja_concurrencia(hoja_concurrencia, proyecto.codigo, metricas_concurrencia)

def escribir_hoja_resumen(hoja, proyecto, metricas_cc, metricas_mi, metricas_bugs_smells=None, metricas_concurrencia=None):
    """Escribe las métricas de un proyecto en la hoja de resumen.

    Args:
        hoja: Hoja de Excel donde se escriben los datos.
        proyecto: Proyecto analizado.
        metricas_cc: Métricas de complejidad.
        metricas_mi: Métricas de mantenibilidad.
        metricas_bugs_smells: Métricas de incidencias, si existen.
        metricas_concurrencia: Métricas de concurrencia, si existen.
    """
    issues_kloc = metricas_bugs_smells.issues_kloc if metricas_bugs_smells else ""
    isi = metricas_bugs_smells.isi if metricas_bugs_smells else ""
    interpretacion_isi = metricas_bugs_smells.interpretacion_isi if metricas_bugs_smells else ""
    valores_resumen = [
        proyecto.codigo,
        proyecto.nombre_proyecto,
        proyecto.herramienta_ia,
        proyecto.modelo_ia,
        proyecto.lenguaje,
        metricas_cc.ccn_promedio,
        metricas_cc.nivel_cc,
        metricas_mi.mi,
        metricas_mi.nivel_mi,
        issues_kloc,
        isi,
        interpretacion_isi,
    ]

    valores_resumen.extend(
        [
            metricas_concurrencia.promedio if metricas_concurrencia else "",
            metricas_concurrencia.interpretacion if metricas_concurrencia else "",
        ]
    )

    escribir_o_actualizar_fila(hoja, proyecto.codigo, valores_resumen)

def escribir_hoja_complejidad(hoja, codigo, metricas_cc):
    """Escribe las métricas de complejidad en la hoja correspondiente.

    Args:
        hoja: Hoja de Excel donde se escriben los datos.
        codigo: Código del proyecto.
        metricas_cc: Métricas de complejidad.
    """
    escribir_o_actualizar_fila(
        hoja,
        codigo,
        [
            codigo,
            metricas_cc.cantidad_funciones,
            metricas_cc.ccn_total,
            metricas_cc.ccn_promedio,
            metricas_cc.nloc_total,
            metricas_cc.nivel_cc,
            metricas_cc.interpretacion_cc,
            metricas_cc.nloc_promedio,
        ],
    )

def escribir_hoja_mantenibilidad(hoja, codigo, metricas_mi):
    """Escribe las métricas de mantenibilidad en la hoja correspondiente.

    Args:
        hoja: Hoja de Excel donde se escriben los datos.
        codigo: Código del proyecto.
        metricas_mi: Métricas de mantenibilidad.
    """
    escribir_o_actualizar_fila(
        hoja,
        codigo,
        [
            codigo,
            metricas_mi.nloc_mi,
            metricas_mi.cantidad_funciones_mi,
            metricas_mi.tokens_codigo,
            metricas_mi.mi,
            metricas_mi.nivel_mi,
            metricas_mi.interpretacion_mi,
        ],
    )

def escribir_hoja_bugs_smells(hoja, proyecto, metricas_bugs_smells):
    """Escribe las métricas de bugs y smells en la hoja correspondiente.

    Args:
        hoja: Hoja de Excel donde se escriben los datos.
        proyecto: Proyecto analizado.
        metricas_bugs_smells: Métricas de bugs y smells.
    """
    if metricas_bugs_smells is None:
        escribir_o_actualizar_fila(hoja, proyecto.codigo, [proyecto.codigo] + [""] * 11)
    else:
        escribir_o_actualizar_fila(
            hoja,
            proyecto.codigo,
            [
                proyecto.codigo,
                metricas_bugs_smells.analizador,
                metricas_bugs_smells.total_issues,
                metricas_bugs_smells.issues_kloc,
                metricas_bugs_smells.isi,
                metricas_bugs_smells.nivel_isi,
                metricas_bugs_smells.interpretacion_isi,
                metricas_bugs_smells.observacion,
                metricas_bugs_smells.cantidad_alta,
                metricas_bugs_smells.cantidad_media,
                metricas_bugs_smells.cantidad_baja,
                formatear_top_reglas(metricas_bugs_smells.top_reglas_violadas),
            ],
        )

        
    fila = buscar_fila_por_codigo(hoja, proyecto.codigo)

    if fila:
        columna_reglas = obtener_mapa_encabezados(hoja)["Reglas incumplidas"]
        celda_top_reglas = hoja.cell(row=fila, column=columna_reglas)

        celda_top_reglas.alignment = Alignment(wrap_text=True, vertical="top")

        hoja.column_dimensions[celda_top_reglas.column_letter].width = 45


def escribir_hoja_concurrencia(hoja, codigo, metricas_concurrencia):
    """Escribe las métricas de concurrencia en la hoja correspondiente.

    Args:
        hoja: Hoja de Excel donde se escriben los datos.
        codigo: Código del proyecto.
        metricas_concurrencia: Métricas de concurrencia.
    """
    if metricas_concurrencia is None:
        escribir_o_actualizar_fila(hoja, codigo, [codigo] + [""] * 6)
    else:
        escribir_o_actualizar_fila(
            hoja,
            codigo,
            [
                codigo,
                metricas_concurrencia.sincronizacion_correcta,
                metricas_concurrencia.ausencia_de_deadlocks,
                metricas_concurrencia.ausencia_de_condicion_de_carrera,
                metricas_concurrencia.uso_correcto_de_exclusion_mutua,
                metricas_concurrencia.promedio,
                metricas_concurrencia.interpretacion,
            ],
        )

def formatear_top_reglas(top_reglas):
    """Convierte el ranking de reglas en texto legible.

    Args:
        top_reglas: Pares formados por regla y cantidad.

    Returns:
        Las reglas y cantidades separadas por punto y coma.
    """
    return "; ".join(f"{regla}: {cantidad}" for regla, cantidad in top_reglas)



def guardar_error_excel(libro, proyecto, mensaje_error):
    """Registra en el libro un error asociado con un proyecto.

    Args:
        libro: Libro de Excel en el que se registra el error.
        proyecto: Proyecto cuyo análisis falló.
        mensaje_error: Descripción del error producido.
    """
    hoja = libro[HOJA_ERRORES]
    hoja.append([proyecto.codigo, proyecto.nombre_proyecto, mensaje_error])


