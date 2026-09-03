"""Evaluación de concurrencia a partir de una rúbrica en Excel."""

import unicodedata
from pathlib import Path

from openpyxl import load_workbook

from util.modelos import MetricaConcurrencia

HOJA_CONCURRENCIA = "Concurrencia"
CAMPOS_RUBRICA = [
    "Sincronización correcta",
    "Ausencia de deadlocks",
    "Ausencia de condición de carrera",
    "Uso correcto de exclusión mutua",
]


def normalizar_texto(texto):
    """Normaliza un valor para realizar comparaciones sin acentos.

    Args:
        texto: Valor que se desea normalizar.

    Returns:
        Texto en minúsculas, sin espacios externos ni diacríticos.
    """
    texto = "" if texto is None else str(texto).strip().lower()
    texto = unicodedata.normalize("NFKD", texto)
    return "".join(c for c in texto if not unicodedata.combining(c))


def cargar_hoja_concurrencia(archivo_datos_entrada):
    """Carga la hoja que contiene la rúbrica de concurrencia.

    Args:
        archivo_datos_entrada: Ruta del libro Excel de entrada.

    Returns:
        La hoja de concurrencia del libro.

    Raises:
        FileNotFoundError: Si el libro no existe.
        ValueError: Si el libro no contiene la hoja esperada.
    """
    ruta = Path(archivo_datos_entrada)
    if not ruta.exists():
        raise FileNotFoundError(f"No se encontró el archivo: {ruta}")
    libro = load_workbook(ruta, data_only=True)
    if HOJA_CONCURRENCIA not in libro.sheetnames:
        raise ValueError(f"No existe la solapa '{HOJA_CONCURRENCIA}' en {ruta}")
    return libro[HOJA_CONCURRENCIA]


def mapear_encabezados(hoja):
    """Crea un mapa normalizado de encabezados a columnas.

    Args:
        hoja: Hoja de cálculo que se desea inspeccionar.

    Returns:
        Diccionario de encabezados normalizados y números de columna.
    """
    encabezados = {}
    for columna in range(1, hoja.max_column + 1):
        valor = hoja.cell(row=1, column=columna).value
        encabezados[normalizar_texto(valor)] = columna
    return encabezados


def obtener_columna(encabezados, nombre):
    """Obtiene la columna asociada con un encabezado obligatorio.

    Args:
        encabezados: Mapa de encabezados a columnas.
        nombre: Nombre del encabezado buscado.

    Returns:
        El número de columna encontrado.

    Raises:
        ValueError: Si el encabezado no existe.
    """
    columna = encabezados.get(normalizar_texto(nombre))
    if columna is None:
        raise ValueError(f"Falta la columna '{nombre}' en la solapa Concurrencia")
    return columna


def buscar_fila_rubrica(hoja, columna_codigo, codigo):
    """Busca la fila de la rúbrica correspondiente a un proyecto.

    Args:
        hoja: Hoja que contiene la rúbrica.
        columna_codigo: Columna con los códigos de proyecto.
        codigo: Código que se desea localizar.

    Returns:
        El número de fila encontrado o ``None``.
    """
    for fila in range(2, hoja.max_row + 1):
        valor = hoja.cell(row=fila, column=columna_codigo).value
        if normalizar_texto(valor) == normalizar_texto(codigo):
            return fila
    return None


def leer_valores_rubrica(hoja, fila, encabezados):
    """Lee los criterios de concurrencia de una fila.

    Args:
        hoja: Hoja que contiene la rúbrica.
        fila: Número de fila que se desea leer.
        encabezados: Mapa de encabezados a columnas.

    Returns:
        Diccionario con los valores de cada criterio.
    """
    valores = {}
    for campo in CAMPOS_RUBRICA:
        columna = obtener_columna(encabezados, campo)
        valores[campo] = hoja.cell(row=fila, column=columna).value
    return valores


def obtener_puntaje(valor, ponderacion):
    """Convierte un nivel textual en su puntaje configurado.

    Args:
        valor: Nivel de concurrencia evaluado.
        ponderacion: Mapa de niveles a puntajes.

    Returns:
        Puntaje asociado con el nivel.

    Raises:
        ValueError: Si el nivel no está contemplado.
    """
    nivel = normalizar_texto(valor).capitalize()
    if nivel not in ponderacion:
        raise ValueError(f"Nivel de concurrencia inválido: {valor}")
    return ponderacion[nivel]


def calcular_puntajes(valores, ponderacion):
    """Calcula los puntajes de todos los criterios de una rúbrica.

    Args:
        valores: Valores textuales de los criterios.
        ponderacion: Mapa de niveles a puntajes.

    Returns:
        Lista de puntajes calculados.
    """
    puntajes = []
    for valor in valores.values():
        puntajes.append(obtener_puntaje(valor, ponderacion))
    return puntajes


def calcular_promedio(puntajes):
    """Calcula el promedio de una colección de puntajes.

    Args:
        puntajes: Puntajes que se desean promediar.

    Returns:
        El promedio redondeado a dos decimales, o cero sin datos.
    """
    if not puntajes:
        return 0
    return round(sum(puntajes) / len(puntajes), 2)


def interpretar_concurrencia(promedio, umbrales):
    """Obtiene la interpretación correspondiente a un promedio.

    Args:
        promedio: Puntaje promedio de concurrencia.
        umbrales: Intervalos de interpretación configurados.

    Returns:
        La interpretación encontrada o un texto sustituto.
    """
    for umbral in umbrales:
        minimo = umbral["min"]
        maximo = umbral["max"]
        if maximo is None and promedio >= minimo:
            return umbral["interpretacion"]
        if maximo is not None and minimo <= promedio <= maximo:
            return umbral["interpretacion"]
    return "Sin interpretación"


def crear_metrica_concurrencia(codigo, valores, ponderacion, umbrales):
    """Construye la métrica de concurrencia de un proyecto.

    Args:
        codigo: Código del proyecto.
        valores: Valores obtenidos de la rúbrica.
        ponderacion: Mapa de niveles a puntajes.
        umbrales: Intervalos de interpretación.

    Returns:
        La métrica de concurrencia calculada.
    """
    puntajes = calcular_puntajes(valores, ponderacion)
    promedio = calcular_promedio(puntajes)
    interpretacion = interpretar_concurrencia(promedio, umbrales)
    return MetricaConcurrencia(
        codigo=codigo,
        sincronizacion_correcta=valores["Sincronización correcta"],
        ausencia_de_deadlocks=valores["Ausencia de deadlocks"],
        ausencia_de_condicion_de_carrera=valores["Ausencia de condición de carrera"],
        uso_correcto_de_exclusion_mutua=valores["Uso correcto de exclusión mutua"],
        promedio=promedio,
        interpretacion=interpretacion,
    )


def analizar_concurrencia(proyecto, archivo_datos_entrada, ponderacion, umbrales, logger):
    """Analiza la rúbrica de concurrencia de un proyecto.

    Args:
        proyecto: Proyecto que se desea evaluar.
        archivo_datos_entrada: Ruta del libro con la rúbrica.
        ponderacion: Mapa de niveles a puntajes.
        umbrales: Intervalos de interpretación.
        logger: Logger de la aplicación.

    Returns:
        La métrica de concurrencia calculada.

    Raises:
        ValueError: Si el proyecto no aparece en la rúbrica.
    """
    logger.info(f"Leyendo rúbrica de concurrencia para {proyecto.codigo}")
    hoja = cargar_hoja_concurrencia(archivo_datos_entrada)
    encabezados = mapear_encabezados(hoja)
    columna_codigo = obtener_columna(encabezados, "Código")
    fila = buscar_fila_rubrica(hoja, columna_codigo, proyecto.codigo)
    if fila is None:
        raise ValueError(f"No se encontró el código '{proyecto.codigo}' en Concurrencia")
    valores = leer_valores_rubrica(hoja, fila, encabezados)
    return crear_metrica_concurrencia(proyecto.codigo, valores, ponderacion, umbrales)
