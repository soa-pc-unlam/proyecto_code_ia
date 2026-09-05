"""Pruebas de escritura de resultados en reportes Excel."""

import unittest

from openpyxl import Workbook

from reportes.excel import crear_hojas_si_no_existen, guardar_error_excel
from modelos.modelos import Proyecto


class ReportesTest(unittest.TestCase):
    """Verifica la creación y actualización de los reportes."""

    def test_los_errores_del_mismo_proyecto_no_se_sobrescriben(self):
        """Conserva todos los errores registrados para un mismo proyecto."""
        libro = Workbook()
        libro.remove(libro.active)
        crear_hojas_si_no_existen(libro)
        proyecto = Proyecto("M1", "Prueba", "ruta", "IA", "Modelo", "Python")
        guardar_error_excel(libro, proyecto, "Primer error")
        guardar_error_excel(libro, proyecto, "Segundo error")
        hoja = libro["Errores"]
        self.assertEqual(hoja.max_row, 3)
        self.assertEqual(hoja.cell(2, 3).value, "Primer error")
        self.assertEqual(hoja.cell(3, 3).value, "Segundo error")


if __name__ == "__main__":
    unittest.main()
