from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment


HOJA_RESUMEN = "Resumen"
HOJA_COMPLEJIDAD = "Complejidad"
HOJA_ERRORES = "Errores"


def crear_o_abrir_excel(archivo_excel):
    if Path(archivo_excel).exists():
        libro = load_workbook(archivo_excel)
    else:
        libro = Workbook()
        libro.remove(libro.active)

    crear_hojas_si_no_existen(libro)
    return libro


def crear_hojas_si_no_existen(libro):
    if HOJA_RESUMEN not in libro.sheetnames:
        hoja = libro.create_sheet(HOJA_RESUMEN)
        hoja.append([
            "Código",
            "Nombre del proyecto",
            "Herramienta IA",
            "Modelo IA",
            "Lenguaje",
            "CCN promedio",
            "Nivel de CC"
        ])

    if HOJA_COMPLEJIDAD not in libro.sheetnames:
        hoja = libro.create_sheet(HOJA_COMPLEJIDAD)
        hoja.append([
            "Código",
            "Cantidad de funciones",
            "CCN Total",
            "CCN promedio",
            "NLOC total",
            "Nivel de CC",
            "Interpretación CC",
            "NLOC promedio"
        ])

    if HOJA_ERRORES not in libro.sheetnames:
        hoja = libro.create_sheet(HOJA_ERRORES)
        hoja.append([
            "Código",
            "Nombre del proyecto",
            "Error"
        ])

    aplicar_estilos_basicos(libro)


def aplicar_estilos_basicos(libro):
    for hoja in libro.worksheets:
        for celda in hoja[1]:
            celda.font = Font(bold=True)
            celda.fill = PatternFill("solid", fgColor="D9EAF7")
            celda.alignment = Alignment(horizontal="center")

        for columna in hoja.columns:
            max_largo = 0
            letra = columna[0].column_letter

            for celda in columna:
                if celda.value is not None:
                    max_largo = max(max_largo, len(str(celda.value)))

            hoja.column_dimensions[letra].width = max_largo + 3


def buscar_fila_por_codigo(hoja, codigo):
    for fila in range(2, hoja.max_row + 1):
        if hoja.cell(row=fila, column=1).value == codigo:
            return fila

    return None


def escribir_o_actualizar_fila(hoja, codigo, valores):
    fila_existente = buscar_fila_por_codigo(hoja, codigo)

    if fila_existente is None:
        hoja.append(valores)
    else:
        for columna, valor in enumerate(valores, start=1):
            hoja.cell(row=fila_existente, column=columna).value = valor


def guardar_resultado_excel(archivo_excel, proyecto, metricas):
    libro = crear_o_abrir_excel(archivo_excel)

    hoja_resumen = libro[HOJA_RESUMEN]
    hoja_complejidad = libro[HOJA_COMPLEJIDAD]

    escribir_o_actualizar_fila(
        hoja_resumen,
        proyecto.codigo,
        [
            proyecto.codigo,
            proyecto.nombre_proyecto,
            proyecto.herramienta_ia,
            proyecto.modelo_ia,
            proyecto.lenguaje,
            metricas.ccn_promedio,
            metricas.nivel_cc
        ]
    )

    escribir_o_actualizar_fila(
        hoja_complejidad,
        proyecto.codigo,
        [
            proyecto.codigo,
            metricas.cantidad_funciones,
            metricas.ccn_total,
            metricas.ccn_promedio,
            metricas.nloc_total,
            metricas.nivel_cc,
            metricas.interpretacion_cc,
            metricas.nloc_promedio
        ]
    )

    aplicar_estilos_basicos(libro)
    generar_graficos(libro)
    libro.save(archivo_excel)


def guardar_error_excel(archivo_excel, proyecto, mensaje_error):
    libro = crear_o_abrir_excel(archivo_excel)
    hoja = libro[HOJA_ERRORES]

    escribir_o_actualizar_fila(
        hoja,
        proyecto.codigo,
        [
            proyecto.codigo,
            proyecto.nombre_proyecto,
            mensaje_error
        ]
    )

    aplicar_estilos_basicos(libro)
    libro.save(archivo_excel)


def generar_graficos(libro):
    nombre_hoja = "Graficos"

    if nombre_hoja in libro.sheetnames:
        del libro[nombre_hoja]

    hoja_graficos = libro.create_sheet(nombre_hoja)
    hoja_resumen = libro[HOJA_RESUMEN]
    hoja_complejidad = libro[HOJA_COMPLEJIDAD]

    if hoja_resumen.max_row < 2:
        return

    grafico_cc = BarChart()
    grafico_cc.title = "CCN promedio por proyecto"
    grafico_cc.y_axis.title = "CCN promedio"
    grafico_cc.x_axis.title = "Proyecto"

    datos_cc = Reference(
        hoja_resumen,
        min_col=6,
        min_row=1,
        max_row=hoja_resumen.max_row
    )

    categorias = Reference(
        hoja_resumen,
        min_col=1,
        min_row=2,
        max_row=hoja_resumen.max_row
    )

    grafico_cc.add_data(datos_cc, titles_from_data=True)
    grafico_cc.set_categories(categorias)

    hoja_graficos.add_chart(grafico_cc, "A1")

    grafico_nloc = BarChart()
    grafico_nloc.title = "NLOC total por proyecto"
    grafico_nloc.y_axis.title = "NLOC total"
    grafico_nloc.x_axis.title = "Proyecto"

    datos_nloc = Reference(
        hoja_complejidad,
        min_col=5,
        min_row=1,
        max_row=hoja_complejidad.max_row
    )

    categorias_nloc = Reference(
        hoja_complejidad,
        min_col=1,
        min_row=2,
        max_row=hoja_complejidad.max_row
    )

    grafico_nloc.add_data(datos_nloc, titles_from_data=True)
    grafico_nloc.set_categories(categorias_nloc)

    hoja_graficos.add_chart(grafico_nloc, "A18")
